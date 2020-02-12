from openpyxl import Workbook
import json

json_file_path = r"C:\Users\Rodzice.Mateusz-PC\Desktop\PythonProjects\ScrapyTutorial2\ZalandoProductsSale\mens-log-10-08.json"

wb = Workbook()
ws = wb.active

with open(json_file_path) as file:
    data = json.loads(file.read())

amount_of_rows = len(data)
amount_of_cols = len(data[0].keys())
column_names = ["url", "gender", "category1", "category2", "category3", "collection", "name", "color", "original_price",
                "real_price", "sale_percent"]
for col in range(1, amount_of_cols + 1):
    c = ws.cell(row=1, column=col, value=column_names[col-1])

for row in range(2, amount_of_rows + 2):
    values = list(data[row - 2].values())
    for col in range(1, amount_of_cols + 1):
        c = ws.cell(row=row, column=col, value=values[col - 1])

wb.save("mens-log-10-08.xlsx")
